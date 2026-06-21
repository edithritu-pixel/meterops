from sqlalchemy import or_
from flask import (
    Flask,
    render_template,
    request,
    redirect,
    jsonify,
    session,
    url_for,
    send_file,
    flash
)
from datetime import datetime
from flask_sqlalchemy import SQLAlchemy
from werkzeug.utils import secure_filename
import pandas as pd
from io import BytesIO
import os

app = Flask(__name__)

# =========================================
# APP CONFIGURATION
# =========================================

app.secret_key = "meter_secret_key"

app.config['SQLALCHEMY_DATABASE_URI'] = 'mysql+pymysql://root:Mwesh2mwesh@localhost/meter_management'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

UPLOAD_FOLDER = 'uploads'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

os.makedirs(UPLOAD_FOLDER, exist_ok=True)

db = SQLAlchemy(app)

# =========================================
# LOGIN DETAILS
# =========================================

USERNAME = "admin"
PASSWORD = "admin123"
ROQ_STATUSES = [

    "New Request",
    "Pending Allocation",
    "Allocated",
    "Awaiting Field Return",
    "Forwarded To Work Order",
    "Pending Batching",
    "At Test Bench",
    "Passed Testing",
    "Failed Testing",
    "Regional Disposal Review",
    "Pending Disposal",
    "HQ Receipt Acknowledged",
    "Closed",
    "Vacate Request",
    "Cancelled",
]

# =========================================
# DATABASE MODELS
# =========================================

class MeterRecord(db.Model):

    id = db.Column(db.Integer, primary_key=True)
    
    transaction_key = db.Column(
        db.String(100),
        unique=True
    )

    account_number = db.Column(db.String(100))
    installation_key = db.Column(db.String(100))
    customer_name = db.Column(db.String(200))
    region = db.Column(
    db.String(100)
)
    allocated_meter = db.Column(db.String(100))
    allocated_to = db.Column(db.String(100))
    field_status = db.Column(db.String(100))
    source = db.Column(db.String(100))
    category = db.Column(db.String(100))
    date_allocated = db.Column(db.Date)
    removed_meter = db.Column(db.String(100))
    final_reading = db.Column(db.String(100))
    return_date = db.Column(db.Date)

    return_date = db.Column(db.Date)

    returned_meter_photo = db.Column(
        db.String(300)
    )

    office_return_document = db.Column(
        db.String(300)
    )

    status = db.Column(db.String(100))
    system_update_status = db.Column(db.String(100))
    meter_photo = db.Column(db.String(200))
    assistant_comment = db.Column(db.String(500))
    pending_reason = db.Column(
    db.String(300)   
)
    remarks = db.Column(
    db.String(500)
)
    operational_remarks = db.Column(
    db.String(1000)
)
    roq_status = db.Column(
        db.String(100),
        default="New Request"
    )

    workflow_stage = db.Column(
    db.String(100),
    default="New Request"
    )

    created_by = db.Column(
    db.String(100)
    )

    date_created = db.Column(db.String(100))
    

    field_remarks = db.Column(
        db.String(500)
    )

    testbench_remarks = db.Column(
    db.String(500)
    )

    disposal_remarks = db.Column(
    db.String(500)
    )

    assigned_team = db.Column(
    db.String(100)
    )

    rts_reason = db.Column(
    db.String(500)
    )

    update_reason = db.Column(
        db.String(500)
    )
    test_bench_result = db.Column(
    db.String(100)
    )

    test_bench_remarks = db.Column(
        db.String(500)
    )
class MeterInventory(db.Model):

    id = db.Column(db.Integer, primary_key=True)

    meter_number = db.Column(
        db.String(100),
        unique=True
    )

    meter_size = db.Column(
    db.String(20)
)
    branch = db.Column(db.String(100))
    region = db.Column(
    db.String(100)
)
    current_location = db.Column(db.String(100))
    status = db.Column(db.String(100))


class CustomerAccount(db.Model):

    id = db.Column(db.Integer, primary_key=True)

    account_number = db.Column(
        db.String(100),
        unique=True
    )

    customer_number = db.Column(db.String(100))

    customer_name = db.Column(db.String(200))

    current_meter = db.Column(db.String(100))

    installation_key = db.Column(
    db.String(100),
    unique=True
    )

    current_meter_installation_date = db.Column(
        db.Date
    )

    current_meter_year_of_manufacture = db.Column(
        db.Integer
    )

    zone = db.Column(db.String(100))
    
    region = db.Column(
    db.String(100)
)

# =========================================
# WORK PLAN MODEL
# =========================================

class WorkPlan(db.Model):

    id = db.Column(db.Integer, primary_key=True)
    
    transaction_key = db.Column(
    db.String(100),
    unique=True
)
    account_number = db.Column(db.String(100))

    customer_name = db.Column(db.String(200))

    region = db.Column(
        db.String(100)
    )

    current_meter = db.Column(db.String(100))

    reported_by = db.Column(db.String(200))
    
    category = db.Column(db.String(100))

    priority = db.Column(db.String(50))

    status = db.Column(db.String(100))

    remarks = db.Column(db.String(500))

    created_at = db.Column(db.DateTime)

    allocated_meter = db.Column(db.String(100))

    allocated_to = db.Column(db.String(100))

    field_status = db.Column(db.String(100))

    date_allocated = db.Column(db.Date)

    date_installed = db.Column(db.Date)

    removed_meter_photo = db.Column(db.String(300))

    office_return_document = db.Column(db.String(300))

    office_return_received = db.Column(db.String(20))

    office_return_date = db.Column(db.Date)

    field_remarks = db.Column(db.String(500))

    work_order_status = db.Column(
    db.String(50),
    default='Pending'
    )

    work_order_comments = db.Column(
        db.Text
    )

    work_order_updated_by = db.Column(
        db.String(100)
    )

    work_order_update_date = db.Column(
        db.String(50)
    )
    

    work_order_comments = db.Column(
        db.Text
    )

    work_order_updated_by = db.Column(
        db.String(100)
    )

    work_order_update_date = db.Column(
        db.String(50)
    )
    cancel_reason = db.Column(
        db.String(500)
    )
class ActivityLog(db.Model):

    id = db.Column(
        db.Integer,
        primary_key=True
    )

    transaction_key = db.Column(
        db.String(100)
    )

    account_number = db.Column(
        db.String(100)
    )

    action = db.Column(
        db.String(200)
    )

    performed_by = db.Column(
        db.String(100)
    )

    remarks = db.Column(
        db.Text
    )

    timestamp = db.Column(
        db.DateTime,
        default=datetime.now
    )



# =========================================
# LOGIN PROTECTION
# =========================================

def login_required():

    if 'user' not in session:
        return False

    return True
from sqlalchemy import or_

def queue_search(workflow_stage):

    search = request.args.get('search', '')

    page = request.args.get(
        'page',
        1,
        type=int
    )

    query = MeterRecord.query.filter_by(
        workflow_stage=workflow_stage
    )

    if search:

        query = query.filter(
            or_(
                MeterRecord.account_number.contains(search),
                MeterRecord.customer_name.contains(search),
                MeterRecord.removed_meter.contains(search),
                MeterRecord.allocated_meter.contains(search)
            )
        )

    pagination = query.paginate(
        page=page,
        per_page=20,
        error_out=False
    )

    return pagination, search

# =========================================
# LOGIN
# =========================================

@app.route('/login', methods=['GET', 'POST'])
def login():

    if request.method == 'POST':
        
        username = request.form['username']
        password = request.form['password']

        if username == USERNAME and password == PASSWORD:

            session['user'] = username

            return redirect(url_for('dashboard'))

        return 'Invalid username or password'
        
        
    return render_template('login.html')

        

# =========================================
# LOGOUT
# =========================================

@app.route('/logout')
def logout():

    session.pop('user', None)

    return redirect(url_for('login'))


# =========================================
# DASHBOARD
# =========================================

@app.route('/dashboard')
@app.route('/')
def dashboard():

    if not login_required():
        return redirect(url_for('login'))

    search = request.args.get('search', '')

    if search:

        records = MeterRecord.query.filter(
            MeterRecord.account_number.contains(search)
        ).all()

    else:

        records = MeterRecord.query.all()

    available_meters = MeterInventory.query.filter_by(
        status='Available'
    ).count()

    pending_installations = MeterRecord.query.filter_by(
        status='Pending Installation'
    ).count()

    completed_meters = MeterRecord.query.filter_by(
        status='Completed'
    ).count()

    pending_updates = MeterRecord.query.filter_by(
        system_update_status='Pending'
    ).count()

    pending_system_updates = MeterRecord.query.filter_by(
    workflow_stage='Awaiting Work Order Update'
    ).count()

    pending_not_updated = MeterRecord.query.filter_by(
    workflow_stage='Not Updated'
    ).count()
    pending_batching = MeterRecord.query.filter_by(
        workflow_stage='Pending Batching'
    ).count()
    batch_ready = pending_batching >= 30

    meters_needed = max(
        0,
        30 - pending_batching
    )
    pending_test_bench = MeterRecord.query.filter_by(
        workflow_stage='Pending Test Bench'
    ).count()

    passed_testing = MeterRecord.query.filter_by(
        workflow_stage='Passed Testing'
    ).count()

    failed_testing = MeterRecord.query.filter_by(
        workflow_stage='Failed Testing'
    ).count()

    pending_disposal = MeterRecord.query.filter_by(
        workflow_stage='Pending Disposal'
    ).count()

    reabsorbed = MeterRecord.query.filter_by(
        workflow_stage='Reabsorbed'
    ).count()
    regional_summary = []
    disposed_meters = MeterInventory.query.filter_by(
        status='Disposed'
    ).count()

    tested_meters = (
        passed_testing +
        failed_testing
)

    regions = [
        'Central',
        'Dagoretti',
        'ISR',
        'Kasarani',
        'Lower Embakasi',
        'Upper Embakasi',
        'Roysambu',
        'Westlands'
    ]

    for region in regions:

        available = MeterInventory.query.filter_by(
            region=region,
            status='Available'
        ).count()

        pending = MeterRecord.query.filter_by(
            region=region,
            workflow_stage='New Request'
        ).count()

        closed_jobs = MeterRecord.query.filter_by(
            region=region,
            field_status='Installed'
        ).count()

        regional_summary.append({
            'region': region,
            'available_meters': available,
            'pending_allocation': pending,
            'closed_jobs': closed_jobs
})
    return render_template(

        'records.html',

        records=records,

        search=search,

        available_meters=available_meters,

        pending_installations=pending_installations,

        completed_meters=completed_meters,

        pending_updates=pending_updates,
    
        pending_system_updates=pending_system_updates,
        
        pending_batching=pending_batching,

        pending_test_bench=pending_test_bench,

        passed_testing=passed_testing,

        failed_testing=failed_testing,

        pending_disposal=pending_disposal,

        reabsorbed=reabsorbed,

        disposed_meters=disposed_meters,

        tested_meters=tested_meters,

        regional_summary=regional_summary,
        
        batch_ready=batch_ready,

        meters_needed=meters_needed,
        )

        

    


# =========================================
# ADD RECORD
# =========================================

@app.route('/add_record', methods=['GET', 'POST'])
def add_record():

    if not login_required():
        return redirect(url_for('login'))

    if request.method == 'POST':

        photo = request.files['meter_photo']
        filename = ""

        photo = request.files['meter_photo']
        filename = ""

        if photo and photo.filename != "":

            filename = secure_filename(photo.filename)

            photo.save(
                os.path.join(
                    app.config['UPLOAD_FOLDER'],
                    filename
                )
            )

        # Prevent duplicate allocation
        existing_allocation = MeterRecord.query.filter_by(
            allocated_meter=request.form['allocated_meter'],
            status='Installed'
        ).first()

        if existing_allocation:
            return "Meter already allocated"

        record = MeterRecord(

            account_number=request.form['account_number'],
            customer_name=request.form['customer_name'],
            allocated_meter=request.form['allocated_meter'],
            allocated_to=request.form['allocated_to'],
            source=request.form['reported_by'],
            category=request.form['category'],
            removed_meter=request.form['removed_meter'],
            meter_photo=filename,
            final_reading=request.form['final_reading'],
            status=request.form['status'],
            assistant_comment=request.form['assistant_comment'],

            system_update_status=(
                'Pending'
                if request.form['status'].strip() == 'Installed'
                else 'Completed'
            )

        )

        db.session.add(record)

        # Update inventory status
        allocated_meter = MeterInventory.query.filter_by(
            meter_number=request.form['allocated_meter']
        ).first()

        if allocated_meter:
            allocated_meter.status = 'Allocated'

        db.session.commit()

        return redirect('/')

    available_meters = MeterInventory.query.filter_by(
        status='Available'
    ).all()
    
    return render_template(
    'add_record.html',
    available_meters=available_meters
)

# =========================================
# EDIT RECORD
# =========================================

@app.route('/edit/<int:id>', methods=['GET', 'POST'])
def edit_record(id):

    if not login_required():
        return redirect(url_for('login'))

    record = MeterRecord.query.get_or_404(id)

    if request.method == 'POST':

        record.allocated_to = request.form['allocated_to']

        record.final_reading = request.form['final_reading']

        record.return_date = request.form['return_date']

        record.status = request.form['status']

        record.assistant_comment = request.form['assistant_comment']
        record.pending_reason = request.form['pending_reason']

        record.operational_remarks = request.form['operational_remarks']

        # FIND INVENTORY METER
        inventory_meter = MeterInventory.query.filter_by(
            meter_number=record.allocated_meter
        ).first()

        # STATUS LOGIC
        if record.status == 'Returned to Store':

            if inventory_meter:
                inventory_meter.status = 'Available'

        else:

            if inventory_meter:
                inventory_meter.status = 'Allocated'

        # SYSTEM UPDATE LOGIC
        if record.status == 'Installed':

            record.system_update_status = 'Pending'

        else:

            record.system_update_status = 'Completed'

        db.session.commit()

        return redirect('/')

    return render_template(
        'edit_record.html',
        record=record
    )

   
# =========================================
# DELETE RECORD
# =========================================

@app.route('/delete/<int:id>')
def delete_record(id):

    if not login_required():
        return redirect(url_for('login'))

    record = MeterRecord.query.get_or_404(id)

    db.session.delete(record)

    db.session.commit()

    return redirect(url_for('dashboard'))
# =========================================
# ROQ PAGE
# =========================================

@app.route('/roq')
def roq():

    if not login_required():
        return redirect(url_for('login'))

    search = request.args.get('search', '')

    page = request.args.get(
        'page',
        1,
        type=int
    )

    query = MeterRecord.query

    if search:

        query = query.filter(
            or_(
                MeterRecord.account_number.contains(search),
                MeterRecord.customer_name.contains(search),
                MeterRecord.allocated_meter.contains(search),
                MeterRecord.removed_meter.contains(search)
            )
        )

    pagination = query.order_by(
        MeterRecord.id.desc()
    ).paginate(
        page=page,
        per_page=20,
        error_out=False
    )
    
    
    return render_template(
        'roq.html',
        records=pagination.items,
        pagination=pagination,
        search=search
    )

# =========================================
# ROQ STATUS FILTER
# =========================================

@app.route('/roq/status/<stage>')
def roq_status(stage):

    if not login_required():
        return redirect(url_for('login'))

    search = request.args.get('search', '')

    page = request.args.get(
        'page',
        1,
        type=int
    )

    query = MeterRecord.query.filter_by(
        workflow_stage=stage
    )

    if search:

        query = query.filter(
            or_(
                MeterRecord.account_number.contains(search),
                MeterRecord.customer_name.contains(search),
                MeterRecord.allocated_meter.contains(search),
                MeterRecord.removed_meter.contains(search)
            )
        )

    pagination = query.order_by(
        MeterRecord.id.desc()
    ).paginate(
        page=page,
        per_page=20,
        error_out=False
    )

    return render_template(
        'roq.html',
        records=pagination.items,
        pagination=pagination,
        search=search,
        current_status=stage
    )
# =========================================
# UPDATE ROQ STATUS
# =========================================

@app.route('/update_roq/<int:id>/<stage>')
def update_roq(id, stage):

    if not login_required():
        return redirect(url_for('login'))

    record = MeterRecord.query.get_or_404(id)

    record.workflow_stage = stage

    db.session.commit()

    return redirect(request.referrer)

@app.route('/activity_logs')
def activity_logs():

    if not login_required():
        return redirect(url_for('login'))

    logs = ActivityLog.query.order_by(
        ActivityLog.timestamp.desc()
    ).all()

    return render_template(
        'activity_logs.html',
        logs=logs
    )
@app.route('/transaction_history/<transaction_key>')
def transaction_history(transaction_key):

    if not login_required():
        return redirect(url_for('login'))

    logs = ActivityLog.query.filter_by(
        transaction_key=transaction_key
    ).order_by(
        ActivityLog.timestamp.desc()
    ).all()

    return render_template(
        'transaction_history.html',
        logs=logs,
        transaction_key=transaction_key
    )    
# =========================================
# INVENTORY
# =========================================

@app.route('/inventory')
def inventory():

    if not login_required():
        return redirect(url_for('login'))

    status = request.args.get('status')

    if status:

        meters = MeterInventory.query.filter_by(
            status=status
        ).all()

    else:

        meters = MeterInventory.query.all()

    return render_template(
        'inventory.html',
        meters=meters
    )


# =========================================
# STOCK SUMMARY
# =========================================

@app.route('/stock_summary')
def stock_summary():

    if not login_required():
        return redirect(url_for('login'))

    meters = MeterInventory.query.filter_by(
        status='Available'
    ).all()

    sizes = sorted(
    list({
        m.meter_size
        for m in meters
        if m.meter_size
    })
)

    summary = {}

    for meter in meters:

        region = meter.region
        size = meter.meter_size

        if region not in summary:

            summary[region] = {
                s: 0 for s in sizes
            }

            summary[region]['Total'] = 0

        if size in summary[region]:

            summary[region][size] += 1

        summary[region]['Total'] += 1

    return render_template(
        'stock_summary.html',
        summary=summary,
        sizes=sizes
    )

# =========================================
# SYSTEM UPDATES
# =========================================

@app.route('/system_updates')
def system_updates():

    if not login_required():
        return redirect(url_for('login'))

    records = MeterRecord.query.filter_by(
        system_update_status='Pending'
    ).all()

    return render_template(
        'system_updates.html',
        records=records
    )

# =========================================
# WORK PLAN
# =========================================
@app.route('/work_plan')
def work_plan():

    if not login_required():
        return redirect(url_for('login'))

    search = request.args.get('search', '')

    query = WorkPlan.query.filter(
        WorkPlan.status.notin_(
            ['Closed', 'Cancelled']
        )
    )

    if search:

        query = query.filter(
            WorkPlan.account_number.contains(search)
        )

    plans = query.all()
    return render_template(
    'work_plan.html',
    plans=plans,
    search=search
)

@app.route('/work_plan/<int:id>', methods=['GET', 'POST'])
def work_plan_detail(id):

    if not login_required():
        return redirect(url_for('login'))

    plan = WorkPlan.query.get_or_404(id)

    available_meters = MeterInventory.query.filter_by(
        status='Available'
    ).all()

    if request.method == 'POST':
        # Allocation
        if request.form.get('allocated_meter'):

            if plan.category == "Vacate":
                flash("Vacate requests do not require meter allocation")
                return redirect(
                    url_for(
                        'work_plan_detail',
                        id=plan.id
                    )
                )

            selected_meter = request.form.get('allocated_meter')
            technician = request.form.get('allocated_to')

            plan.allocated_meter = selected_meter
            plan.allocated_to = technician
            plan.date_allocated = datetime.now().strftime('%Y-%m-%d')

            log = ActivityLog(
                transaction_key=plan.transaction_key,
                account_number=plan.account_number,
                action="Meter Allocated",
                performed_by=session.get('user'),
                remarks=f"Allocated {selected_meter} to {technician}"
            )

            db.session.add(log)

            # Update ROQ record
            record = MeterRecord.query.filter_by(
                transaction_key=plan.transaction_key
            ).first()

            if record:
                record.allocated_meter = selected_meter
                record.allocated_to = technician
                record.workflow_stage = 'Allocated'

            inventory_meter = MeterInventory.query.filter_by(
                meter_number=selected_meter
            ).first()

            if inventory_meter:
                inventory_meter.status = 'Allocated'

            db.session.commit()

            return redirect(
                url_for(
                    'work_plan_detail',
                    id=plan.id
                )
            )
        
        # Field Status

        elif request.form.get('field_status'):

    
            field_status = request.form.get(
                'field_status'
            )
            print("FIELD STATUS =", field_status)

            plan.field_status = field_status
            
            removed_photo = request.files.get(
                'removed_meter_photo'
            )

            if removed_photo and removed_photo.filename:

                filename = secure_filename(
                    removed_photo.filename
                )

                removed_photo.save(
                    os.path.join(
                        app.config['UPLOAD_FOLDER'],
                        filename
                    )
                )

                plan.removed_meter_photo = filename


            return_doc = request.files.get(
                'office_return_document'
            )

            if return_doc and return_doc.filename:

                doc_filename = secure_filename(
                    return_doc.filename
                )

                return_doc.save(
                    os.path.join(
                        app.config['UPLOAD_FOLDER'],
                        doc_filename
                    )
                )

                plan.office_return_document = doc_filename
                
            plan.field_remarks = request.form.get(
                'field_remarks'
            )

            plan.date_installed = datetime.now().strftime(
                '%Y-%m-%d'
            )

            record = MeterRecord.query.filter_by(
                transaction_key=plan.transaction_key
            ).first()

            if record:

                record.field_status = field_status

                record.field_remarks = request.form.get(
                    'field_remarks'
                )

                record.returned_meter_photo = (
                    plan.removed_meter_photo
                )

                record.office_return_document = (
                    plan.office_return_document
                )
                            
                if field_status == 'Installed':

                    inventory_meter = MeterInventory.query.filter_by(
                        meter_number=plan.allocated_meter
                    ).first()

                    if inventory_meter:
                        inventory_meter.status = 'Installed'

                    record.workflow_stage = 'Awaiting Work Order Update'

                    if plan.category == "Replacement":
                        record.removed_meter = plan.current_meter

                    elif plan.category == "Stolen Meter":
                        record.removed_meter = "STOLEN"

                    elif plan.category == "New Connection":
                        record.removed_meter = None


                elif field_status == 'Returned':

                    record.workflow_stage = 'Awaiting Work Order Update'

                    if plan.category == "Vacate":

                        record.removed_meter = plan.current_meter

                        returned_meter = MeterInventory.query.filter_by(
                            meter_number=plan.current_meter
                        ).first()

                        if returned_meter:
                            returned_meter.status = 'Pending Test Bench'
                            returned_meter.region = record.region
                            returned_meter.current_location = 'Store'

                        else:
                            returned_meter = MeterInventory(
                                meter_number=plan.current_meter,
                                region=record.region,
                                current_location='Store',
                                status='Pending Test Bench'
                            )
                            plan.field_status = field_status
                            plan.field_remarks = request.form.get('field_remarks')
                            plan.office_return_received = True
                            plan.office_return_date = datetime.now().strftime('%Y-%m-%d')

                            db.session.commit()

                            return redirect(
                                url_for(
                                    'work_plan_detail',
                                    id=plan.id
                                )
                            )


                            db.session.add(returned_meter)
                    inventory_meter = MeterInventory.query.filter_by(
                        meter_number=plan.allocated_meter
                    ).first()

                    if inventory_meter:
                        inventory_meter.status = 'Installed'

                    record.workflow_stage = 'Awaiting Work Order Update'

                    # CATEGORY RULES
                    if plan.category == "Replacement":
                        record.removed_meter = plan.current_meter

                        returned_meter = MeterInventory.query.filter_by(
                            meter_number=plan.current_meter
                        ).first()

                        if returned_meter:
                            returned_meter.status = 'Pending Test Bench'
                            returned_meter.region = record.region
                            returned_meter.current_location = 'Store'

                        else:
                            returned_meter = MeterInventory(
                                meter_number=plan.current_meter,
                                region=record.region,
                                current_location='Store',
                                status='Pending Test Bench'
                            )
                            db.session.add(returned_meter)

                    elif plan.category == "Stolen Meter":
                        record.removed_meter = "STOLEN"

                    elif plan.category == "New Connection":
                        record.removed_meter = None

                    elif plan.category == "Vacate":
                        record.removed_meter = plan.current_meter

                        returned_meter = MeterInventory.query.filter_by(
                            meter_number=plan.current_meter
                        ).first()

                        if returned_meter:
                            returned_meter.status = 'Pending Test Bench'
                            returned_meter.region = record.region
                            returned_meter.current_location = 'Store'

                        else:
                            returned_meter = MeterInventory(
                                meter_number=plan.current_meter,
                                region=record.region,
                                current_location='Store',
                                status='Pending Test Bench'
                            )
                            db.session.add(returned_meter)
                            print("SETTING REMARKS")
                        print(record.remarks)
                        inventory_meter = (
                            MeterInventory.query.filter_by(
                                meter_number=plan.allocated_meter
                            ).first()
                        )

                    if inventory_meter:

                        inventory_meter.status = (
                            'Available'
                        )
                        log = ActivityLog(
                            transaction_key=plan.transaction_key,
                            account_number=plan.account_number,
                            action="Field Status Updated",
                            performed_by=session.get('user'),
                            remarks=f"Status changed to {plan.field_status}"
                        )

                        db.session.add(log)
                        db.session.commit()

                return redirect(
                    url_for(
                        'work_plan_detail',
                        id=plan.id
                    )
                )    
               # Work Order
        elif request.form.get('work_order_submit'):

            work_order_status = request.form.get(
                'work_order_status'
            )
            

            plan.work_order_status = work_order_status
            plan.work_order_status = request.form.get(
                'work_order_status'
            )
            plan.work_order_comments = request.form.get(
                'work_order_comments'
            )
            log = ActivityLog(
                transaction_key=plan.transaction_key,
                account_number=plan.account_number,
                action="Work Order Updated",
                performed_by=session.get('user'),
                remarks=f"Work order marked as {work_order_status}"
            )

            db.session.add(log)
            plan.work_order_update_date = datetime.now().strftime(
                '%Y-%m-%d'
            )

            record = MeterRecord.query.filter_by(
                transaction_key=plan.transaction_key
            ).first()

            if record:

                
                if work_order_status == 'Updated':

                    # CATEGORY ROUTING
                    if plan.category in ['Replacement', 'Vacate']:
                        record.workflow_stage = 'Pending Batching'
                        plan.status = 'Closed'

                    elif plan.category in ['Stolen Meter', 'New Connection']:
                        record.workflow_stage = 'Closed'
                        record.status = 'Completed'
                        plan.status = 'Closed'

                    record.pending_reason = None
                    
                else:

                    record.workflow_stage = (
                        'Not Updated'
                    )
                    record.pending_reason = (
                        plan.work_order_comments
                    )

                    
                    record.pending_reason = (
                        request.form.get(
                            'work_order_comments'
                        )
                    )

                    

            db.session.commit()

            return redirect(
                url_for(
                    'work_plan_detail',
                    id=plan.id
                )
            )

    return render_template(
        'work_plan_detail.html',
        plan=plan,
        available_meters=available_meters
    )
# =========================================
# ADD WORK PLAN ITEM
# =========================================


    return render_template('add_work_plan.html')
# =========================================
# UPLOAD METERS
# =========================================
@app.route('/add_work_plan', methods=['GET', 'POST'])
def add_work_plan():

    if not login_required():
        return redirect(url_for('login'))

    if request.method == 'POST':

        txn_key = f"TXN-{datetime.now().strftime('%Y%m%d%H%M%S')}"

        plan = WorkPlan(
            transaction_key=txn_key,
            account_number=request.form['account_number'],
            customer_name=request.form['customer_name'],
            region=request.form['region'],
            current_meter=request.form['current_meter'],
            reported_by=request.form['reported_by'],
            category=request.form['category'],
            priority=request.form['priority'],
            status=request.form['status'],
            remarks=request.form['remarks'],
            created_at=request.form['created_at']
        )

        # CREATE ROQ RECORD
        record = MeterRecord(
            transaction_key=txn_key,
            installation_key=request.form['installation_key'],
            account_number=request.form['account_number'],
            customer_name=request.form['customer_name'],
            category=request.form['category'],
            region=request.form.get('region'),
            source=request.form['reported_by'],
            remarks=request.form['remarks'],
            status='Pending',
            workflow_stage='New Request'
        )

        db.session.add(plan)
        db.session.add(record)

        log = ActivityLog(
            transaction_key=txn_key,
            account_number=plan.account_number,
            action="Work Plan Created",
            performed_by=session.get('user'),
            remarks="Initial work plan created"
        )

        db.session.add(log)
        db.session.commit()

        return redirect('/work_plan')

    return render_template('add_work_plan.html')



# =========================================
# UPLOAD CUSTOMERS
# =========================================

@app.route('/upload_customers', methods=['GET', 'POST'])
def upload_customers():

    if not login_required():
        return redirect(url_for('login'))

    if request.method == 'POST':

        file = request.files['file']

        filepath = os.path.join(
            app.config['UPLOAD_FOLDER'],
            file.filename
        )

        file.save(filepath)

        df = pd.read_csv(filepath)

        for _, row in df.iterrows():

            account_number = str(
                row['account_number']
            ).strip()

            if account_number == 'nan':
                continue

            existing = CustomerAccount.query.filter_by(
                account_number=account_number
            ).first()

            if not existing:

                customer = CustomerAccount(
    account_number=account_number,

    customer_name=str(
        row['customer_name']
    ),

    current_meter=str(
        row['current_meter']
    ),

    customer_number=str(
        row['customer_number']
    ),

    zone=str(
        row['zone']
    ),

    region=str(
        row['region']
    ),
    installation_key=row['installation_key'],
    current_meter_installation_date=row['current_meter_installation_date'],
    current_meter_year_of_manufacture=row['current_meter_year_of_manufacture'],

)
                
                db.session.add(customer)

        db.session.commit()

        return redirect('/')

    return render_template(
        'upload_customers.html'
    )


# =========================================
# AUTO POPULATE CUSTOMER
# =========================================

@app.route('/get_customer/<account_number>')
def get_customer(account_number):

    if not login_required():
        return jsonify({})

    customer = CustomerAccount.query.filter_by(
        account_number=account_number
    ).first()

    if customer:

        return jsonify({

    'customer_name': customer.customer_name,

    'customer_number': customer.customer_number,

    'current_meter': customer.current_meter,

    'installation_key': customer.installation_key,
    'current_meter_installation_date': customer.current_meter_installation_date,
    'current_meter_year_of_manufacture': customer.current_meter_year_of_manufacture,
    'zone': customer.zone,
    'region': customer.region

})

    return jsonify({

    'customer_name': '',

    'customer_number': '',

    'current_meter': '',


    'zone': '',
    'region': ''

})


# =========================================
# EXPORT EXCEL
# =========================================

@app.route('/export')
def export_excel():

    if not login_required():
        return redirect(url_for('login'))

    records = MeterRecord.query.all()

    data = []

    for r in records:

        data.append({
            'Account Number': r.account_number,
            'Customer Name': r.customer_name,
            'Allocated Meter': r.allocated_meter,
            'Allocated To': r.allocated_to,
            'Source': r.source,
            'Status': r.status,
            'Removed Meter': r.removed_meter,
            'Final Reading': r.final_reading,
            'Assistant Comment': r.assistant_comment
        })

    df = pd.DataFrame(data)

    export_path = 'meter_records.xlsx'

    df.to_excel(export_path, index=False)

    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill

    wb = load_workbook(export_path)
    ws = wb.active

    # Header styling
    header_fill = PatternFill(
        start_color="1F2937",
        end_color="1F2937",
        fill_type="solid"
    )

    for cell in ws[1]:

        cell.font = Font(
            bold=True,
            color="FFFFFF"
        )

        cell.fill = header_fill

    # Auto width
    for column in ws.columns:

        max_length = 0
        column_letter = column[0].column_letter

        for cell in column:

            try:

                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))

            except:
                pass

        adjusted_width = max_length + 5

        ws.column_dimensions[
            column_letter
        ].width = adjusted_width

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(export_path)

    return send_file(
        export_path,
        as_attachment=True
    )


# =========================================
# CREATE DATABASE
# =========================================

with app.app_context():
    db.create_all()


@app.route('/allocate/<int:id>', methods=['GET', 'POST'])
def allocate_meter(id):

    record = MeterRecord.query.get_or_404(id)

    if request.method == 'POST':

        record.allocated_meter = request.form.get(
            'allocated_meter'
        )

        record.allocated_to = request.form.get(
            'allocated_to'
        )

        record.workflow_stage = "Allocated"

        db.session.commit()

        flash('Meter allocated successfully')

        return redirect(url_for('roq'))

    available_meters = MeterInventory.query.filter(
    MeterInventory.status.in_([
        'Available',
        'Reabsorbed'
    ])
).all()
    return render_template(
        'allocate_meter.html',
        record=record,
        available_meters=available_meters
    )
@app.route('/rts-close/<int:id>', methods=['GET', 'POST'])
def rts_close(id):

    record = MeterRecord.query.get_or_404(id)

    if request.method == 'POST':

        record.rts_reason = request.form.get(
            'reason'
        )

        record.workflow_stage = "RTS Closed"

        

        db.session.commit()

        flash('Record closed as RTS')

        return redirect(url_for('roq'))

    return render_template(
        'rts_close.html',
        record=record
    )


@app.route('/forward-workorder/<int:id>')
def forward_workorder(id):

    record = MeterRecord.query.get_or_404(id)

    record.workflow_stage = "Forwarded To Work Order"

    db.session.commit()

    flash('Forwarded to work order')

    return redirect(url_for('roq'))
    
@app.route('/mark-updated/<int:id>')
def mark_updated(id):

    record = MeterRecord.query.get_or_404(id)

    record.workflow_stage = "Updated"

    db.session.commit()

    flash('System updated successfully')

    return redirect(url_for('roq'))
@app.route('/not-updated/<int:id>', methods=['GET', 'POST'])
def not_updated(id):

    record = MeterRecord.query.get_or_404(id)

    if request.method == 'POST':

        record.update_reason = request.form.get(
            'reason'
        )

        record.workflow_stage = "Not Updated"

        db.session.commit()

        flash('Marked as not updated')

        return redirect(url_for('roq'))

    return render_template(
        'not_updated.html',
        record=record
    )

# =========================================
# RUN APP
# =========================================
@app.route('/uploads/<filename>')
def uploaded_file(filename):

    return send_file(
        os.path.join(
            app.config['UPLOAD_FOLDER'],
            filename
        )
    )
@app.route('/metering_officer')
def metering_officer():

    if not login_required():
        return redirect(url_for('login'))

    pagination, search = queue_search(
        'Pending Metering Officer'
    )

    return render_template(
        'metering_officer.html',
        records=pagination.items,
        pagination=pagination,
        search=search
    )
    
@app.route(
    '/metering_officer/<int:id>',
    methods=['GET', 'POST']
)
def metering_officer_detail(id):

    if not login_required():
        return redirect(url_for('login'))

    record = MeterRecord.query.get_or_404(id)

    if request.method == 'POST':

        record.operational_remarks = request.form.get(
            'metering_remarks'
        )

        record.workflow_stage = (
            'Pending Batching'
        )

        db.session.commit()

        flash(
            'Meter forwarded to batching'
        )

        return redirect(
            url_for('metering_officer')
        )

    return render_template(
        'metering_officer_detail.html',
        record=record
    )
@app.route('/batching_queue')
def batching_queue():

    if not login_required():
        return redirect(url_for('login'))

    pagination, search = queue_search(
        'Pending Batching'
    )
    pending_batching = MeterRecord.query.filter_by(
        workflow_stage='Pending Batching'
    ).count()
    return render_template(
    'batching_queue.html',
    records=pagination.items,
    pagination=pagination,
    search=search,
    pending_batching=pending_batching
)
@app.route('/create_batch')
def create_batch():

    if not login_required():
        return redirect(url_for('login'))

    records = MeterRecord.query.filter_by(
        workflow_stage='Pending Batching'
    ).limit(30).all()

    for record in records:

        record.workflow_stage = (
            'Pending Test Bench'
        )

        inventory_meter = MeterInventory.query.filter_by(
            meter_number=record.removed_meter
        ).first()

        if inventory_meter:
            inventory_meter.status = 'Test Bench'

            inventory_meter = MeterInventory.query.filter_by(
                meter_number=record.removed_meter
            ).first()

            if inventory_meter:
                inventory_meter.status = 'Test Bench'
    db.session.commit()

    flash(
        f'{len(records)} meters forwarded to Test Bench'
    )

    return redirect(
        url_for('batching_queue')
    )

@app.route(
    '/batching/<int:id>',
    methods=['GET', 'POST']
)
def batching_detail(id):

    if not login_required():
        return redirect(url_for('login'))

    record = MeterRecord.query.get_or_404(id)

    if request.method == 'POST':

        record.workflow_stage = (
            'Pending Test Bench'
        )

        db.session.commit()

        flash(
            'Forwarded to Test Bench'
        )

        return redirect(
            url_for('batching_queue')
        )

    return render_template(
        'batching_detail.html',
        record=record
    ) 
@app.route(
    '/forward_selected_batch',
    methods=['POST']
)
def forward_selected_batch():

    if not login_required():
        return redirect(url_for('login'))

    selected = request.form.getlist(
        'selected_records'
    )

    if not selected:

        flash(
            'No meters selected'
        )

        return redirect(
            url_for('batching_queue')
        )

    records = MeterRecord.query.filter(
        MeterRecord.id.in_(selected)
    ).all()

    for record in records:

        record.workflow_stage = (
            'Pending Test Bench'
        )

    db.session.commit()

    flash(
        f'{len(records)} meters forwarded'
    )

    return redirect(
        url_for('batching_queue')
    )
@app.route('/test_bench')
def test_bench():

    if not login_required():
        return redirect(url_for('login'))

    pagination, search = queue_search(
        'Pending Test Bench'
    )

    return render_template(
        'test_bench.html',
        records=pagination.items,
        pagination=pagination,
        search=search
    )    
@app.route(
    '/test_bench/<int:id>',
    methods=['GET', 'POST']
)
def test_bench_detail(id):

    if not login_required():
        return redirect(url_for('login'))

    record = MeterRecord.query.get_or_404(id)

    if request.method == 'POST':

        record.test_bench_result = request.form.get(
            'test_bench_result'
        )

        record.test_bench_remarks = request.form.get(
            'test_bench_remarks'
        )

        if record.test_bench_result == 'Pass':

            record.workflow_stage = 'Passed Testing'

            inventory_meter = MeterInventory.query.filter_by(
                meter_number=record.removed_meter
            ).first()

            if inventory_meter:
                inventory_meter.status = 'Reabsorbed'

        else:

            record.workflow_stage = 'Failed Testing'

            inventory_meter = MeterInventory.query.filter_by(
                meter_number=record.removed_meter
            ).first()

            if inventory_meter:
                inventory_meter.status = 'Faulty'

        log = ActivityLog(
            transaction_key=record.transaction_key,
            account_number=record.account_number,
            action='Bench Test Completed',
            performed_by=session.get('user'),
            remarks=f"Result: {record.test_bench_result}"
        )

        db.session.add(log)

        db.session.commit()

        flash(
            'Test bench results saved'
        )

        page = request.args.get('page', 1)

        return redirect(
            url_for(
                'test_bench',
                page=page
            )
        )

    return render_template(
        'test_bench_detail.html',
        record=record
    )
@app.route('/passed_testing')
def passed_testing():

    if not login_required():
        return redirect(url_for('login'))

    pagination, search = queue_search(
        'Passed Testing'
    )

    return render_template(
        'passed_testing.html',
        records=pagination.items,
        pagination=pagination,
        search=search
    )
@app.route(
    '/passed_testing/<int:id>',
    methods=['GET', 'POST']
)
def passed_testing_detail(id):

    if not login_required():
        return redirect(url_for('login'))

    record = MeterRecord.query.get_or_404(id)

    if request.method == 'POST':

        record.workflow_stage = 'Reabsorbed'

        inventory_meter = MeterInventory.query.filter_by(
            meter_number=record.removed_meter
        ).first()

        if inventory_meter:

            inventory_meter.status = 'Reabsorbed'

        else:

            inventory_meter = MeterInventory(
            meter_number=record.removed_meter,
            region=record.region,
            current_location='Store',
            status='Reabsorbed'
        )

            db.session.add(inventory_meter)
            log = ActivityLog(
                transaction_key=record.transaction_key,
                account_number=record.account_number,
                action='Meter Reabsorbed',
                performed_by=session.get('user'),
                remarks=f"Meter {record.removed_meter} reabsorbed into inventory"
            )

            db.session.add(log)
        db.session.commit()

        flash('Meter reabsorbed into inventory')

        return redirect(
            url_for('passed_testing')
        )

    return render_template(
        'passed_testing_detail.html',
        record=record
    )
@app.route('/pending_disposal')
def pending_disposal():

    if not login_required():
        return redirect(url_for('login'))

    pagination, search = queue_search(
        'Pending Disposal'
    )

    return render_template(
        'pending_disposal.html',
        records=pagination.items,
        pagination=pagination,
        search=search
    )
@app.route(
    '/pending_disposal/<int:id>',
    methods=['GET', 'POST']
)
def pending_disposal_detail(id):

    if not login_required():
        return redirect(url_for('login'))

    record = MeterRecord.query.get_or_404(id)

    if request.method == 'POST':

        record.disposal_remarks = request.form.get(
            'disposal_remarks'
        )

        record.workflow_stage = (
    'Disposed'
)

        inventory_meter = MeterInventory.query.filter_by(
            meter_number=record.removed_meter
        ).first()

        if inventory_meter:
            inventory_meter.status = 'Disposed'
        log = ActivityLog(
            transaction_key=record.transaction_key,
            account_number=record.account_number,
            action='Meter Disposed',
            performed_by=session.get('user'),
            remarks=record.disposal_remarks
        )

        db.session.add(log)
        db.session.commit()

        flash(
             'Meter received and disposed successfully'
)

        return redirect(
            url_for('pending_disposal')
        )

    return render_template(
        'pending_disposal_detail.html',
        record=record
    )



@app.route('/dispose/<int:id>', methods=['POST'])
def dispose_meter(id):

    if not login_required():
        return redirect(url_for('login'))

    record = MeterRecord.query.get_or_404(id)

    record.workflow_stage = (
        'Disposed'
    )

    inventory_meter = MeterInventory.query.filter_by(
        meter_number=record.removed_meter
    ).first()

    if inventory_meter:
        inventory_meter.status = 'Disposed'

    db.session.commit()

    flash(
    'Meter disposed successfully'
)


    return redirect(
        url_for('pending_disposal')
    )
@app.route('/failed_testing')
def failed_testing():

    if not login_required():
        return redirect(url_for('login'))

    pagination, search = queue_search(
        'Failed Testing'
    )

    return render_template(
        'failed_testing.html',
        records=pagination.items,
        pagination=pagination,
        search=search
    )

@app.route(
    '/failed_testing/<int:id>',
    methods=['GET', 'POST']
)
def failed_testing_detail(id):

    if not login_required():
        return redirect(url_for('login'))

    record = MeterRecord.query.get_or_404(id)

    if request.method == 'POST':

        record.disposal_remarks = request.form.get(
            'disposal_remarks'
        )

        record.workflow_stage = (
            'Regional Disposal Review'
        )
        log = ActivityLog(
            transaction_key=record.transaction_key,
            account_number=record.account_number,
            action='Failed Testing',
            performed_by=session.get('user'),
            remarks='Meter failed bench testing and forwarded for disposal'
        )

        db.session.add(log)
        db.session.commit()

        flash(
            'Forwarded to disposal'
        )
        
        return redirect(
            url_for('failed_testing')
        )

    return render_template(
        'failed_testing_detail.html',
        record=record
    )

# =========================================
# REPORTS
# =========================================

@app.route('/reports')
def reports():

    if not login_required():
        return redirect(url_for('login'))

    return render_template('reports.html')


@app.route('/reports/meter_replacement')
def meter_replacement_report():

    if not login_required():
        return redirect(url_for('login'))

    query = MeterRecord.query.filter(
        MeterRecord.field_status == 'Installed'
    )

    query = apply_report_filters(query)

    pagination = report_pagination(query)

    return render_template(
        'meter_replacement_report.html',
        records=pagination.items,
        pagination=pagination
    )
@app.route('/reports/meter_replacement/export')
def export_meter_replacement_report():

    if not login_required():
        return redirect(url_for('login'))

    query = MeterRecord.query.filter(
        MeterRecord.field_status == 'Installed'
    )

    query = apply_report_filters(query)

    records = query.all()

    data = []

    for record in records:

        data.append({
            'Account Number': record.account_number,
            'Customer Name': record.customer_name,
            'Region': record.region,
            'Removed Meter': record.removed_meter,
            'Allocated Meter': record.allocated_meter,
            'Field Status': record.field_status,
            'Workflow Stage': record.workflow_stage
        })

    df = pd.DataFrame(data)

    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    if start_date and end_date:

        filename = (
            f'meter_replacement_report_'
            f'{start_date}_to_{end_date}.xlsx'
        )

    else:

        filename = (
            f'meter_replacement_report_'
            f'{datetime.now().strftime("%d%m%Y")}.xlsx'
        )

    output = BytesIO()

    with pd.ExcelWriter(
        output,
        engine='openpyxl'
    ) as writer:

        df.to_excel(
            writer,
            index=False,
            sheet_name='Meter Replacement'
        )

    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name=filename
    )
@app.route('/reports/workflow_summary')
def workflow_summary_report():

    if not login_required():
        return redirect(url_for('login'))

    query = MeterRecord.query

    query = apply_report_filters(query)

    records = query.all()

    stages = {}

    for record in records:

        stage = record.workflow_stage

        if not stage:
            continue

        if stage not in stages:
            stages[stage] = 0

        stages[stage] += 1

    return render_template(
        'workflow_summary_report.html',
        stages=stages
    )
@app.route('/reports/batching')
def batching_report():

    if not login_required():
        return redirect(url_for('login'))

    pagination = report_pagination(
        query = MeterRecord.query.filter(
            MeterRecord.workflow_stage == 'Pending Batching'
        )
    )

    return render_template(
    'batching_report.html',
    records=pagination.items,
    pagination=pagination
 
)
@app.route('/reports/passed_testing')
def passed_testing_report():

    if not login_required():
        return redirect(url_for('login'))

    query = MeterRecord.query.filter(
        MeterRecord.workflow_stage == 'Passed Testing'
    )

    query = apply_report_filters(query)

    pagination = report_pagination(query)

    return render_template(
    'passed_testing_report.html',
    records=pagination.items,
    pagination=pagination
)
@app.route('/reports/failed_testing')
def failed_testing_report():

    if not login_required():
        return redirect(url_for('login'))

    page = request.args.get(
    'page',
    1,
    type=int
    )

    query = MeterRecord.query.filter(
        MeterRecord.workflow_stage == 'Failed Testing'
    )

    query = apply_report_filters(query)

    pagination = report_pagination(query)

    return render_template(
    'failed_testing_report.html',
    records=pagination.items,
    pagination=pagination
) 
@app.route('/reports/failed_testing/export')
def export_failed_testing_report():

    if not login_required():
        return redirect(url_for('login'))

    query = MeterRecord.query.filter(
        MeterRecord.workflow_stage == 'Failed Testing'
    )

    query = apply_report_filters(query)

    records = query.all()

    data = []

    for record in records:

        data.append({
            'Account Number': record.account_number,
            'Customer Name': record.customer_name,
            'Removed Meter': record.removed_meter,
            'Allocated Meter': record.allocated_meter,
            'Workflow Stage': record.workflow_stage
        })

    df = pd.DataFrame(data)

    today = datetime.now().strftime('%d%m%Y')

    filename = (
        f'failed_testing_report_{today}.xlsx'
    )

    output = BytesIO()

    with pd.ExcelWriter(
        output,
        engine='openpyxl'
    ) as writer:

        df.to_excel(
            writer,
            index=False,
            sheet_name='Failed Testing'
        )

    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name=filename
    )
@app.route('/reports/passed_testing/export')
def export_passed_testing_report():

    if not login_required():
        return redirect(url_for('login'))

    query = MeterRecord.query.filter(
        MeterRecord.workflow_stage == 'Passed Testing'
    )

    query = apply_report_filters(query)

    records = query.all()

    data = []

    for record in records:

        data.append({
            'Account Number': record.account_number,
            'Customer Name': record.customer_name,
            'Removed Meter': record.removed_meter,
            'Allocated Meter': record.allocated_meter,
            'Workflow Stage': record.workflow_stage
        })

    df = pd.DataFrame(data)

    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    if start_date and end_date:

        filename = (
            f'passed_testing_report_'
            f'{start_date}_to_{end_date}.xlsx'
        )

    else:

        filename = (
            f'passed_testing_report_'
            f'{datetime.now().strftime("%d%m%Y")}.xlsx'
        )

    output = BytesIO()

    with pd.ExcelWriter(
        output,
        engine='openpyxl'
    ) as writer:

        df.to_excel(
            writer,
            index=False,
            sheet_name='Passed Testing'
        )

    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name=filename
    )

@app.route('/reports/disposal')
def disposal_report():

    if not login_required():
        return redirect(url_for('login'))

    query = MeterRecord.query.filter(
        MeterRecord.workflow_stage.in_([
            'Pending Disposal',
            'Disposed'
        ])
    )

    query = apply_report_filters(query)

    pagination = report_pagination(query)

    return render_template(
        'disposal_report.html',
        records=pagination.items,
        pagination=pagination
    )

@app.route('/reports/batching/export')
def export_batching_report():

    if not login_required():
        return redirect(url_for('login'))

    query = MeterRecord.query.filter(
        MeterRecord.workflow_stage == 'Pending Batching'
    )

    query = apply_report_filters(query)

    records = query.all()

    data = []

    for record in records:

        data.append({
            'Account Number': record.account_number,
            'Customer Name': record.customer_name,
            'Region': record.region,
            'Allocated Meter': record.allocated_meter,
            'Removed Meter': record.removed_meter,
            'Workflow Stage': record.workflow_stage
        })

    df = pd.DataFrame(data)

    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    if start_date and end_date:

        filename = (
            f'batching_report_'
            f'{start_date}_to_{end_date}.xlsx'
        )

    else:

        filename = (
            f'batching_report_'
            f'{datetime.now().strftime("%d%m%Y")}.xlsx'
        )

    output = BytesIO()

    with pd.ExcelWriter(
        output,
        engine='openpyxl'
    ) as writer:

        df.to_excel(
            writer,
            index=False,
            sheet_name='Batching'
        )

    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name=filename
    )
@app.route('/reports/technician_performance')
def technician_performance_report():

    if not login_required():
        return redirect(url_for('login'))

    query = MeterRecord.query.filter(
        MeterRecord.allocated_to.isnot(None)
    )

    query = apply_report_filters(query)

    records = query.all()

    technicians = {}

    for record in records:

        tech = record.allocated_to

        if not tech:
            continue

        if tech not in technicians:
            technicians[tech] = 0

        technicians[tech] += 1

    return render_template(
        'technician_performance_report.html',
        technicians=technicians
    )
@app.route('/regional_disposal')
def regional_disposal():

    if not login_required():
        return redirect(url_for('login'))

    pagination, search = queue_search(
        'Regional Disposal Review'
    )

    return render_template(
        'regional_disposal.html',
        records=pagination.items,
        pagination=pagination,
        search=search
    )
@app.route(
    '/regional_disposal/<int:id>',
    methods=['GET', 'POST']
)
def regional_disposal_detail(id):

    if not login_required():
        return redirect(url_for('login'))

    record = MeterRecord.query.get_or_404(id)

    if request.method == 'POST':

        record.workflow_stage = (
            'Pending Disposal'
        )
        inventory_meter = MeterInventory.query.filter_by(
            meter_number=record.removed_meter
        ).first()

        
        db.session.commit()

        flash(
            'Forwarded to HQ Disposal'
        )

        return redirect(
            url_for('regional_disposal')
        )

    return render_template(
        'regional_disposal_detail.html',
        record=record
    )
@app.route('/hq_disposal')
def hq_disposal():

    pagination, search = queue_search(
        'Disposed'
    )

    return render_template(
        'hq_disposal.html',
        records=pagination.items,
        pagination=pagination,
        search=search
    )  
@app.route('/reports/disposal/export')
def export_disposal_report():

    if not login_required():
        return redirect(url_for('login'))

    query = MeterRecord.query.filter(
        MeterRecord.workflow_stage.in_([
            'Pending Disposal',
            'Disposed'
        ])
    )

    query = apply_report_filters(query)

    records = query.all()

    data = []

    for record in records:

        data.append({
            'Account Number': record.account_number,
            'Customer Name': record.customer_name,
            'Removed Meter': record.removed_meter,
            'Allocated Meter': record.allocated_meter,
            'Workflow Stage': record.workflow_stage
        })

    df = pd.DataFrame(data)

    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    if start_date and end_date:

        filename = (
            f'disposal_report_{start_date}_to_{end_date}.xlsx'
        )

    else:

        filename = (
            f'disposal_report_{datetime.now().strftime("%d%m%Y")}.xlsx'
        )

    output = BytesIO()

    with pd.ExcelWriter(
        output,
        engine='openpyxl'
    ) as writer:

        df.to_excel(
            writer,
            index=False,
            sheet_name='Disposal'
        )

    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name=filename
    )
# =========================================
# REPORT HELPERS
# =========================================

def report_pagination(query):

    page = request.args.get(
        'page',
        1,
        type=int
    )

    return query.paginate(
        page=page,
        per_page=20,
        error_out=False
    )   
def get_report_filters():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    region = request.args.get('region')

    return start_date, end_date, region

def apply_report_filters(query):

    query = apply_date_filter(query)

    query = apply_region_filter(query)

    return query 

def apply_region_filter(query):

    region = request.args.get('region')

    if region:

        query = query.filter(
            MeterRecord.region == region
        )

    return query
def apply_date_filter(query):

    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    if start_date:
        query = query.filter(
            MeterRecord.date_created >= start_date
        )

    if end_date:
        query = query.filter(
            MeterRecord.date_created <= end_date
        )

    return query
if __name__ == '__main__':
    app.run(
        host='0.0.0.0',
        port=8000,
        debug=True
    )

